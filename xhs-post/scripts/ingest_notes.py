#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""把爆文库 xlsx 转成结构化的 templates.json + 单篇全文 txt。

新增参考笔记时重跑本脚本即可，不要手工改 templates.json。
用法: python ingest_notes.py <xlsx路径> [输出目录]
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

COLS = ["tpl_id", "tpl_name", "img_category", "title", "url",
        "body_struct", "img_struct", "full_text", "engagement", "tier"]

# 养号轨：不背转化 KPI，其 tier 无参考意义
NURTURE_VOICES = {"conflict-story", "meme-remix"}


def parse_engagement(s):
    """'点赞156，收藏62，评论30' -> dict。缺失或格式异常返回 None。"""
    if not s:
        return None
    nums = {}
    for key, label in (("likes", "点赞"), ("saves", "收藏"), ("comments", "评论")):
        m = re.search(rf"{label}\s*(\d+)", str(s))
        if m:
            nums[key] = int(m.group(1))
    return nums or None


def main():
    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else "xhs_examples.xlsx")
    outdir = Path(sys.argv[2] if len(sys.argv) > 2 else Path(__file__).parents[2] / "data")
    notes_dir = outdir / "notes"
    notes_dir.mkdir(parents=True, exist_ok=True)

    labels_path = outdir / "labels.json"
    labels = {}
    if labels_path.exists():
        labels = json.loads(labels_path.read_text(encoding="utf-8")).get("labels", {})

    ws = openpyxl.load_workbook(xlsx, data_only=True).worksheets[0]
    records = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 11)]
        if not any(vals):
            continue
        rec = dict(zip(COLS, vals))
        rec["row"] = r
        rec["engagement"] = parse_engagement(rec["engagement"])
        rec["char_count"] = len(str(rec["full_text"] or ""))
        rec["tags"] = re.findall(r"#([^\s#]+)", str(rec["full_text"] or ""))
        # 三轴标注来自 data/labels.json，与 xlsx 解耦
        rec.update(labels.get(str(r), {"voice": None, "placement": None,
                                       "body_carries": None, "art": None}))
        rec["tier"] = str(rec["tier"] or "").strip() or None   # S/A/B 转化分级
        rec["track"] = "nurture" if rec.get("voice") in NURTURE_VOICES else "conversion"
        if rec["track"] == "nurture":
            rec["tier"] = None       # 养号轨不背转化 KPI，tier 不可比
        rec["conversion"] = None     # 具体转化数值，拿到后回填
        records.append(rec)

    for rec in records:
        slug = f"{rec['row']:03d}-{re.sub(r'[^\w一-鿿]+', '_', str(rec['title']))[:40]}"
        rec["note_file"] = f"notes/{slug}.txt"
        (notes_dir / f"{slug}.txt").write_text(str(rec["full_text"] or ""), encoding="utf-8")

    index = [{k: v for k, v in rec.items() if k != "full_text"} for rec in records]
    (outdir / "templates.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"{len(records)} 篇 -> {outdir / 'templates.json'}")
    unlabeled = [r["row"] for r in records if not r["body_struct"]]
    if unlabeled:
        print(f"未标注正文结构的行: {unlabeled}")


if __name__ == "__main__":
    main()
